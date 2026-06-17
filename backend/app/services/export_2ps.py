"""Generate ПС / ЛС Excel export matching template (Форма 1ПС + Форма 2ПС per section)."""
from __future__ import annotations

import io
from datetime import datetime
from math import floor
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Styles ────────────────────────────────────────────────────────────────────

_THIN = Side(style="thin")
_MED  = Side(style="medium")
_NONE = Side(style=None)


def _b(*args):
    """_b(left, right, top, bottom) or _b() for all-thin."""
    if not args:
        return Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
    l, r, t, bo = args
    return Border(left=l, right=r, top=t, bottom=bo)


_B_ALL   = _b()
_B_MED   = _b(_MED, _MED, _MED, _MED)
_B_NONE  = Border()

_F_MAIN  = Font(name="Times New Roman", size=10)
_F_BOLD  = Font(name="Times New Roman", size=10, bold=True)
_F_SM    = Font(name="Times New Roman", size=9)
_F_H1    = Font(name="Times New Roman", size=12, bold=True)
_F_H2    = Font(name="Times New Roman", size=11, bold=True)

_FILL_H  = PatternFill(fill_type="solid", fgColor="D9D9D9")
_FILL_T  = PatternFill(fill_type="solid", fgColor="F2F2F2")
_FILL_SEC = PatternFill(fill_type="solid", fgColor="D6E4F7")

_AC  = Alignment(horizontal="center",  vertical="center", wrap_text=True)
_AR  = Alignment(horizontal="right",   vertical="center", wrap_text=False)
_AL  = Alignment(horizontal="left",    vertical="center", wrap_text=True)
_AWT = Alignment(horizontal="left",    vertical="top",    wrap_text=True)
_ARC = Alignment(horizontal="right",   vertical="center", wrap_text=True)

_RUB_FMT = '# ##0.00'
_THO_FMT = '# ##0.0000'   # тысячи руб., 4 дес. знака как в шаблоне


def _set(ws, row, col, value, font=None, align=None, border=None, fill=None, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:   c.font           = font
    if align:  c.alignment      = align
    if border: c.border         = border
    if fill:   c.fill           = fill
    if fmt:    c.number_format  = fmt
    return c


def _mg(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


# ── Word helper ───────────────────────────────────────────────────────────────

def _rub_words(amount: float) -> str:
    """Amount in rubles → Russian words string."""
    total = floor(amount)
    kopecks = round((amount - total) * 100)

    units   = ["", "один", "два", "три", "четыре", "пять", "шесть", "семь", "восемь", "девять",
               "десять", "одиннадцать", "двенадцать", "тринадцать", "четырнадцать", "пятнадцать",
               "шестнадцать", "семнадцать", "восемнадцать", "девятнадцать"]
    units_f = ["", "одна", "две", "три", "четыре", "пять", "шесть", "семь", "восемь", "девять",
               "десять", "одиннадцать", "двенадцать", "тринадцать", "четырнадцать", "пятнадцать",
               "шестнадцать", "семнадцать", "восемнадцать", "девятнадцать"]
    tens    = ["", "десять", "двадцать", "тридцать", "сорок", "пятьдесят",
               "шестьдесят", "семьдесят", "восемьдесят", "девяносто"]
    hunds   = ["", "сто", "двести", "триста", "четыреста", "пятьсот",
               "шестьсот", "семьсот", "восемьсот", "девятьсот"]

    def _chunk(n, feminine=False):
        u = units_f if feminine else units
        parts = []
        if n // 100:
            parts.append(hunds[n // 100])
        r = n % 100
        if r < 20:
            if r: parts.append(u[r])
        else:
            parts.append(tens[r // 10])
            if r % 10: parts.append(u[r % 10])
        return " ".join(parts)

    def _suffix(n, m1, m24, m5):
        r10, r100 = n % 10, n % 100
        if r10 == 1 and r100 != 11: return m1
        if 2 <= r10 <= 4 and not (11 <= r100 <= 14): return m24
        return m5

    parts = []
    billions = total // 1_000_000_000; total %= 1_000_000_000
    millions = total // 1_000_000;     total %= 1_000_000
    thousands = total // 1_000;         remainder = total % 1_000

    if billions:
        parts.append(f"{_chunk(billions)} {_suffix(billions, 'миллиард', 'миллиарда', 'миллиардов')}")
    if millions:
        parts.append(f"{_chunk(millions)} {_suffix(millions, 'миллион', 'миллиона', 'миллионов')}")
    if thousands:
        parts.append(f"{_chunk(thousands, feminine=True)} {_suffix(thousands, 'тысяча', 'тысячи', 'тысяч')}")
    if remainder:
        parts.append(_chunk(remainder))

    w = " ".join(parts) if parts else "ноль"
    w = w[0].upper() + w[1:]
    return f"{w} рублей {kopecks:02d} копеек"


# ── ЛС detail sheet ───────────────────────────────────────────────────────────
# Columns: A=№, B:E=Наименование, F:G=Ссылка, H=Расчёт, I=Стоимость (тыс.)

_LS_NCOLS = 9
_LS_WIDTHS = [6, 12, 10, 8, 10, 20, 16, 28, 16]   # A..I


def _write_ls_sheet(ws, ls_num: str, project_name: str, stage: str,
                    section_name: str, positions: list[dict],
                    vat_rate: float, quarter: str):
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True

    for i, w in enumerate(_LS_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1

    # R1: system info + form reference top-right
    _mg(ws, r, 1, r, 7)
    _set(ws, r, 1, "", font=_F_SM, align=_AL)
    _mg(ws, r, 8, r, 9)
    _set(ws, r, 8,
         "Приложение №7\nк приказу МС №707/пр\nот 01.09.2021 г.",
         font=_F_SM, align=_ARC)
    ws.row_dimensions[r].height = 40
    r += 1

    # R2: empty
    ws.row_dimensions[r].height = 6
    r += 1

    # R3: Смета № ЛС-XX
    _mg(ws, r, 1, r, 9)
    _set(ws, r, 1, f"Смета № {ls_num}", font=_F_H1, align=_AC)
    ws.row_dimensions[r].height = 20
    r += 1

    # R4: на проектные работы
    _mg(ws, r, 1, r, 9)
    _set(ws, r, 1, "на проектные работы", font=_F_MAIN, align=_AC)
    ws.row_dimensions[r].height = 16
    r += 1

    # R5: section name (наименование стройки / объекта)
    _mg(ws, r, 1, r, 9)
    _set(ws, r, 1, section_name or project_name, font=_F_MAIN, align=_AC)
    ws.row_dimensions[r].height = 18
    r += 1

    # R6: (наименование стройки)
    _mg(ws, r, 1, r, 9)
    _set(ws, r, 1, "(наименование стройки)", font=_F_SM,
         align=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[r].height = 14
    r += 1

    info_pairs = [
        ("Заказчик", ""),
        ("", "(наименование организации)"),
        ("Проектная организация", "ООО «Интеллект-Строй»"),
        ("", "(наименование организации)"),
        ("Составлена в уровне цен на", quarter or datetime.now().strftime("%I кв. %Y г.")),
    ]
    for label, val in info_pairs:
        _mg(ws, r, 1, r, 2)
        _set(ws, r, 1, label, font=_F_MAIN if label else _F_SM, align=_AL)
        _mg(ws, r, 3, r, 9)
        _set(ws, r, 3, val,   font=_F_MAIN if label else _F_SM, align=_AL)
        ws.row_dimensions[r].height = 16
        r += 1

    # Empty
    ws.row_dimensions[r].height = 6
    r += 1

    # Header row (single row, no vertical span)
    _set(ws, r, 1, "№\nп/п",
         font=_F_BOLD, align=_AC, border=_B_ALL, fill=_FILL_H)
    _mg(ws, r, 2, r, 5)
    _set(ws, r, 2, "Наименование объекта проектирования или вида работы",
         font=_F_BOLD, align=_AC, border=_B_ALL, fill=_FILL_H)
    _mg(ws, r, 6, r, 7)
    _set(ws, r, 6, "Наименование, номера глав, таблиц, параграфов\nнормативных документов",
         font=_F_BOLD, align=_AC, border=_B_ALL, fill=_FILL_H)
    _set(ws, r, 8, "Расчет стоимости",
         font=_F_BOLD, align=_AC, border=_B_ALL, fill=_FILL_H)
    _set(ws, r, 9, "Сметная стоимость,\nтыс. руб.",
         font=_F_BOLD, align=_AC, border=_B_ALL, fill=_FILL_H)
    ws.row_dimensions[r].height = 40
    r += 1

    # Number row
    for ci, txt in [(1, "1"), (2, "2"), (6, "3"), (8, "4"), (9, "5")]:
        _set(ws, r, ci, txt, font=_F_SM, align=_AC, border=_B_ALL, fill=_FILL_H)
    _mg(ws, r, 2, r, 5)
    _mg(ws, r, 6, r, 7)
    ws.row_dimensions[r].height = 14
    r += 1

    # ── Positions ──────────────────────────────────────────────────────────
    for pos in positions:
        cost_tys = pos["cost"] / 1000   # convert руб → тыс. руб.

        # Col A: №
        _set(ws, r, 1, pos["num"], font=_F_MAIN, align=_AC, border=_B_ALL)
        # Col B-E: name
        _mg(ws, r, 2, r, 5)
        _set(ws, r, 2, pos["name"], font=_F_MAIN, align=_AWT, border=_B_ALL)
        # Col F-G: justification
        _mg(ws, r, 6, r, 7)
        _set(ws, r, 6, pos["justification"], font=_F_SM, align=_AWT, border=_B_ALL)
        # Col H: formula
        _set(ws, r, 8, pos["formula"], font=_F_SM, align=_AWT, border=_B_ALL)
        # Col I: cost (тыс. руб.)
        _set(ws, r, 9, cost_tys, font=_F_BOLD, align=_AR, border=_B_ALL, fmt=_THO_FMT)
        ws.row_dimensions[r].height = 48
        r += 1

    # ── Summary ────────────────────────────────────────────────────────────
    total_cost    = sum(p["cost"] for p in positions)
    vat_amount    = round(total_cost * vat_rate / 100, 2)
    total_with_vat = round(total_cost + vat_amount, 2)

    cost_tys      = total_cost / 1000
    vat_tys       = vat_amount / 1000
    total_tys     = total_with_vat / 1000

    def _sum(label, val, bold=False, fill=None):
        nonlocal r
        _mg(ws, r, 1, r, 8)
        _set(ws, r, 1, label, font=_F_BOLD if bold else _F_MAIN,
             align=_AL, border=_B_ALL, fill=fill)
        _set(ws, r, 9, val, font=_F_BOLD if bold else _F_MAIN,
             align=_AR, border=_B_ALL, fill=fill, fmt=_THO_FMT)
        ws.row_dimensions[r].height = 18
        r += 1

    _sum("Итого без учета НДС", cost_tys, bold=True, fill=_FILL_T)

    _mg(ws, r, 1, r, 1); _set(ws, r, 1, "1", font=_F_MAIN, align=_AC, border=_B_ALL)
    _mg(ws, r, 2, r, 8); _set(ws, r, 2, "Итого без НДС", font=_F_MAIN, align=_AL, border=_B_ALL)
    _set(ws, r, 9, cost_tys, font=_F_MAIN, align=_AR, border=_B_ALL, fmt=_THO_FMT)
    ws.row_dimensions[r].height = 16; r += 1

    _mg(ws, r, 1, r, 1); _set(ws, r, 1, "2", font=_F_MAIN, align=_AC, border=_B_ALL)
    _mg(ws, r, 2, r, 7); _set(ws, r, 2, "Налог на добавленную стоимость (НДС)",
                                font=_F_MAIN, align=_AL, border=_B_ALL)
    _set(ws, r, 8, f"{vat_rate:.0f} %", font=_F_MAIN, align=_AC, border=_B_ALL)
    _set(ws, r, 9, vat_tys, font=_F_MAIN, align=_AR, border=_B_ALL, fmt=_THO_FMT)
    ws.row_dimensions[r].height = 16; r += 1

    _mg(ws, r, 1, r, 1); _set(ws, r, 1, "3", font=_F_MAIN, align=_AC, border=_B_ALL)
    _mg(ws, r, 2, r, 8); _set(ws, r, 2, "Итого по смете", font=_F_BOLD, align=_AL,
                                border=_B_ALL, fill=_FILL_T)
    _set(ws, r, 9, total_tys, font=_F_BOLD, align=_AR, border=_B_ALL, fill=_FILL_T, fmt=_THO_FMT)
    ws.row_dimensions[r].height = 16; r += 1

    # Signature
    r += 1
    _mg(ws, r, 1, r, 9)
    words = _rub_words(total_with_vat)
    _set(ws, r, 1,
         f"Итого по смете: {total_with_vat:,.2f} руб. ({words})",
         font=_F_MAIN, align=_AWT)
    ws.row_dimensions[r].height = 28


# ── ПС summary sheet ──────────────────────────────────────────────────────────
# Columns: A=№, B-C=Перечень работ, D=Характеристики, E=Ссылка,
#          F-G=Изыскат.(0), H-I=Проектных, J=Всего

_PS_NCOLS = 10
_PS_WIDTHS = [6, 22, 16, 20, 12, 8, 8, 8, 14, 14]  # A..J


def _write_ps_sheet(ws, project_name: str, stage: str,
                    result: dict, ls_info: list[dict], vat_rate: float):
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True

    for i, w in enumerate(_PS_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    quarter = result.get("price_index_period", datetime.now().strftime("%I кв. %Y г."))
    year    = datetime.now().year
    r = 1

    # R1: Приложение + Форма 1ПС
    _mg(ws, r, 1, r, 8)
    _set(ws, r, 1,
         f"Приложение № 1  к договору №_____ от __.__.{year} г.",
         font=_F_MAIN, align=_AL)
    _mg(ws, r, 9, r, 10)
    _set(ws, r, 9, "Форма 1ПС", font=_F_BOLD,
         align=Alignment(horizontal="right", vertical="center"))
    ws.row_dimensions[r].height = 18; r += 1

    # R2: empty
    ws.row_dimensions[r].height = 6; r += 1

    # R3: title
    _mg(ws, r, 1, r, 10)
    _set(ws, r, 1, "Смета ПС", font=_F_H1, align=_AC)
    ws.row_dimensions[r].height = 22; r += 1

    _mg(ws, r, 1, r, 10)
    _set(ws, r, 1, "на проектные (изыскательские) работы", font=_F_MAIN, align=_AC)
    ws.row_dimensions[r].height = 16; r += 1

    info_rows = [
        ("Наименование строительства",                      project_name),
        ("Стадии проектирования",                           stage),
        ("Наименование проектной организации - генпроект.", "ООО «Интеллект-Строй»"),
        ("Составлена в ценах",                              quarter),
    ]
    for label, val in info_rows:
        _mg(ws, r, 1, r, 2)
        _set(ws, r, 1, label, font=_F_MAIN, align=_AL, border=_B_ALL)
        _mg(ws, r, 3, r, 10)
        _set(ws, r, 3, val,   font=_F_MAIN, align=_AL, border=_B_ALL)
        ws.row_dimensions[r].height = 24; r += 1

    ws.row_dimensions[r].height = 6; r += 1

    # ── Table header ─────────────────────────────────────────────────────
    # Row 1 of header: main captions
    _set(ws, r, 1,  "№\nп.п.",                          font=_F_BOLD, align=_AC, border=_B_ALL, fill=_FILL_H)
    _mg(ws, r, 2,  r, 3)
    _set(ws, r, 2,  "Перечень выполняемых работ",        font=_F_BOLD, align=_AC, border=_B_ALL, fill=_FILL_H)
    _set(ws, r, 4,  "Характеристики\nпроектируемого объекта",
                                                         font=_F_BOLD, align=_AC, border=_B_ALL, fill=_FILL_H)
    _set(ws, r, 5,  "Ссылка на\n№ сметы",               font=_F_BOLD, align=_AC, border=_B_ALL, fill=_FILL_H)
    _mg(ws, r, 6,  r, 10)
    _set(ws, r, 6,  "Стоимость работ, тыс. руб.",        font=_F_BOLD, align=_AC, border=_B_ALL, fill=_FILL_H)
    ws.row_dimensions[r].height = 32; r += 1

    # Row 2 of header: sub-captions for cost columns
    _set(ws, r, 1,  "",  font=_F_SM, align=_AC, border=_B_ALL, fill=_FILL_H)
    _mg(ws, r, 2,  r, 3)
    _set(ws, r, 2,  "",  font=_F_SM, align=_AC, border=_B_ALL, fill=_FILL_H)
    _set(ws, r, 4,  "",  font=_F_SM, align=_AC, border=_B_ALL, fill=_FILL_H)
    _set(ws, r, 5,  "",  font=_F_SM, align=_AC, border=_B_ALL, fill=_FILL_H)
    _mg(ws, r, 6,  r, 8)
    _set(ws, r, 6,  "Изыскательских", font=_F_BOLD, align=_AC, border=_B_ALL, fill=_FILL_H)
    _mg(ws, r, 9,  r, 9)
    _set(ws, r, 9,  "Проектных",       font=_F_BOLD, align=_AC, border=_B_ALL, fill=_FILL_H)
    _set(ws, r, 10, "Всего",           font=_F_BOLD, align=_AC, border=_B_ALL, fill=_FILL_H)
    ws.row_dimensions[r].height = 20; r += 1

    # ── Section rows ──────────────────────────────────────────────────────
    for i, ls in enumerate(ls_info, 1):
        cost_no_vat = ls["cost"]
        cost_tys    = cost_no_vat / 1000

        _set(ws, r, 1, i,               font=_F_MAIN, align=_AC,  border=_B_ALL)
        _mg(ws, r, 2, r, 3)
        _set(ws, r, 2, ls["name"],       font=_F_MAIN, align=_AWT, border=_B_ALL)
        _set(ws, r, 4, "",               font=_F_MAIN, align=_AL,  border=_B_ALL)
        _set(ws, r, 5, ls["sheet"],      font=_F_MAIN, align=_AC,  border=_B_ALL)
        _mg(ws, r, 6, r, 8)
        _set(ws, r, 6, 0,                font=_F_MAIN, align=_AR,  border=_B_ALL, fmt=_THO_FMT)
        _set(ws, r, 9, cost_tys,         font=_F_MAIN, align=_AR,  border=_B_ALL, fmt=_THO_FMT)
        _set(ws, r, 10, cost_tys,        font=_F_MAIN, align=_AR,  border=_B_ALL, fmt=_THO_FMT)
        ws.row_dimensions[r].height = 20; r += 1

    # Итого row
    total_cost_no_vat = sum(ls["cost"] for ls in ls_info)
    total_vat         = round(total_cost_no_vat * vat_rate / 100, 2)
    total_with_vat    = round(total_cost_no_vat + total_vat, 2)
    total_tys         = total_cost_no_vat / 1000
    vat_tys           = total_vat / 1000
    grand_tys         = total_with_vat / 1000

    _mg(ws, r, 1, r, 4)
    _set(ws, r, 1, "Итого:", font=_F_BOLD, align=_AL, border=_B_ALL, fill=_FILL_T)
    _set(ws, r, 5, "",       font=_F_BOLD, align=_AL, border=_B_ALL, fill=_FILL_T)
    _mg(ws, r, 6, r, 8)
    _set(ws, r, 6, 0,        font=_F_BOLD, align=_AR, border=_B_ALL, fill=_FILL_T, fmt=_THO_FMT)
    _set(ws, r, 9, total_tys, font=_F_BOLD, align=_AR, border=_B_ALL, fill=_FILL_T, fmt=_THO_FMT)
    _set(ws, r, 10, total_tys, font=_F_BOLD, align=_AR, border=_B_ALL, fill=_FILL_T, fmt=_THO_FMT)
    ws.row_dimensions[r].height = 18; r += 1

    ws.row_dimensions[r].height = 8; r += 1  # gap

    # НДС summary block
    def _ps_sum(num, label, val_col8, val_col9, val_col10, bold=False, fill=None):
        nonlocal r
        _set(ws, r, 1, num, font=_F_MAIN, align=_AC, border=_B_ALL)
        _mg(ws, r, 2, r, 7)
        _set(ws, r, 2, label, font=_F_BOLD if bold else _F_MAIN, align=_AL, border=_B_ALL, fill=fill)
        _set(ws, r, 8, val_col8, font=_F_BOLD if bold else _F_MAIN, align=_AC, border=_B_ALL, fill=fill)
        _set(ws, r, 9, val_col9, font=_F_BOLD if bold else _F_MAIN, align=_AR, border=_B_ALL,
             fill=fill, fmt=_THO_FMT if isinstance(val_col9, float) else None)
        _set(ws, r, 10, val_col10, font=_F_BOLD if bold else _F_MAIN, align=_AR, border=_B_ALL,
             fill=fill, fmt=_THO_FMT if isinstance(val_col10, float) else None)
        ws.row_dimensions[r].height = 18; r += 1

    _ps_sum("1", "Итого без НДС",             "",               total_tys, total_tys)
    _ps_sum("2", "НДС (НДС)",                 f"{vat_rate:.0f}%", vat_tys,   vat_tys)
    _ps_sum("3", "Итого по смете",             "",               grand_tys, grand_tys,
            bold=True, fill=_FILL_T)

    ws.row_dimensions[r].height = 8; r += 1

    # Total in words
    _mg(ws, r, 1, r, 10)
    words = _rub_words(total_with_vat)
    _set(ws, r, 1,
         f"Итого по смете: {total_with_vat:,.2f} руб. ({words})",
         font=_F_MAIN, align=_AWT)
    ws.row_dimensions[r].height = 32; r += 1

    # Signature area
    r += 1
    _mg(ws, r, 1, r, 4)
    _set(ws, r, 1, "Подрядчик: ____________________", font=_F_MAIN, align=_AL)
    _mg(ws, r, 7, r, 10)
    _set(ws, r, 7, "Заказчик: ____________________", font=_F_MAIN, align=_AL)
    ws.row_dimensions[r].height = 20


# ── Main entry point ──────────────────────────────────────────────────────────

def generate_2ps_excel(project_name: str, stage: str, result: dict[str, Any]) -> bytes:
    positions = result.get("positions", [])
    vat_rate  = float(result.get("vat_rate", 22))
    quarter   = result.get("price_index_period", "")

    # Group by section_num (0 = no section)
    sections: dict[int, dict] = {}
    for pos in positions:
        snum  = int(pos.get("section_num") or 0)
        sname = pos.get("section_name") or ""
        if snum not in sections:
            sections[snum] = {"name": sname, "positions": []}
        sections[snum]["positions"].append(pos)

    section_keys = sorted(sections.keys())

    wb = openpyxl.Workbook()

    ls_info: list[dict] = []

    for i, snum in enumerate(section_keys, 1):
        sec = sections[snum]
        ls_name = f"ЛС-{i:02d}"
        ws = wb.create_sheet(title=ls_name)
        sec_positions = sec["positions"]
        sec_cost     = sum(p["cost"] for p in sec_positions)
        sec_vat      = round(sec_cost * vat_rate / 100, 2)
        sec_total    = round(sec_cost + sec_vat, 2)

        _write_ls_sheet(
            ws, ls_name, project_name, stage,
            sec["name"] or project_name,
            sec_positions, vat_rate, quarter,
        )
        ls_info.append({
            "sheet": ls_name,
            "name":  sec["name"] or project_name,
            "cost":  sec_cost,
            "vat":   sec_vat,
            "total": sec_total,
        })

    # ПС summary sheet always first
    ps_ws = wb.create_sheet(title="ПС", index=0)
    _write_ps_sheet(ps_ws, project_name, stage, result, ls_info, vat_rate)

    # Remove default empty sheet
    for name in ["Sheet", "Лист", "Лист1"]:
        if name in wb.sheetnames:
            del wb[name]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
