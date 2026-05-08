import asyncio
import io
import json
import os
import shutil
import uuid
from typing import AsyncGenerator, List

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.api.deps import require_admin
from app.config import settings
from app.database import get_db
from app.models import AuditLog, ReferenceBook, ReferenceRow, User
from app.schemas import ReferenceBookOut, ReferenceBookUpdate

router = APIRouter(prefix="/admin/references", tags=["admin"])


@router.get("", response_model=List[ReferenceBookOut])
def list_references(db: Session = Depends(get_db), _: User = Depends(require_admin)):
    return db.query(ReferenceBook).order_by(ReferenceBook.code, ReferenceBook.version.desc()).all()


@router.post("", response_model=ReferenceBookOut, status_code=201)
def upload_reference(
    code: str,
    official_name: str,
    file: UploadFile = File(...),
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=422, detail="Допустимый формат: PDF")

    # Determine next version for this code
    last = db.query(ReferenceBook).filter(ReferenceBook.code == code).order_by(ReferenceBook.version.desc()).first()
    next_version = (last.version + 1) if last else 1

    dest_dir = os.path.join(settings.uploads_dir, "references", code)
    os.makedirs(dest_dir, exist_ok=True)
    filename = f"v{next_version}_{uuid.uuid4().hex}.pdf"
    dest_path = os.path.join(dest_dir, filename)
    with open(dest_path, "wb") as f:
        shutil.copyfileobj(file.file, f)

    book = ReferenceBook(
        code=code,
        official_name=official_name,
        version=next_version,
        pdf_filename=file.filename,
        pdf_path=dest_path,
        status="requires_parsing",
    )
    db.add(book)
    db.flush()
    db.add(AuditLog(
        user_id=current_user.id, action="upload_reference",
        resource_type="reference_book", resource_id=book.id,
        details={"code": code, "version": next_version},
    ))
    db.commit()
    db.refresh(book)
    return book


@router.delete("/{book_id}", status_code=204)
def delete_reference(
    book_id: int,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    book = _get_book(book_id, db)
    if book.is_active:
        raise HTTPException(status_code=422, detail="Нельзя удалить активный справочник")
    pdf_path = book.pdf_path
    db.query(ReferenceRow).filter(ReferenceRow.book_version_id == book_id).delete()
    db.add(AuditLog(
        user_id=current_user.id, action="delete_reference",
        resource_type="reference_book", resource_id=book.id,
        details={"code": book.code, "version": book.version},
    ))
    db.delete(book)
    db.commit()
    if pdf_path and os.path.exists(pdf_path):
        os.remove(pdf_path)


@router.get("/{book_id}", response_model=ReferenceBookOut)
def get_reference(book_id: int, db: Session = Depends(get_db), _: User = Depends(require_admin)):
    book = _get_book(book_id, db)
    return book


@router.patch("/{book_id}", response_model=ReferenceBookOut)
def update_reference(
    book_id: int, body: ReferenceBookUpdate,
    db: Session = Depends(get_db), _: User = Depends(require_admin),
):
    book = _get_book(book_id, db)
    if body.notes is not None:
        book.notes = body.notes
    if body.parse_prompt is not None:
        book.parse_prompt = body.parse_prompt
    db.commit()
    db.refresh(book)
    return book


@router.post("/{book_id}/activate", response_model=ReferenceBookOut)
def activate_reference(
    book_id: int,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    book = _get_book(book_id, db)
    if book.status != "consistent":
        raise HTTPException(status_code=422, detail="Активировать можно только валидированный справочник (consistent)")

    # Deactivate previous active version of same code
    db.query(ReferenceBook).filter(
        ReferenceBook.code == book.code,
        ReferenceBook.is_active == True,
        ReferenceBook.id != book.id,
    ).update({"is_active": False, "status": "archived"})

    from datetime import datetime
    book.is_active = True
    book.activated_at = datetime.utcnow()
    db.add(AuditLog(
        user_id=current_user.id, action="activate_reference",
        resource_type="reference_book", resource_id=book.id,
    ))
    db.commit()
    db.refresh(book)
    return book


@router.post("/{book_id}/rollback", response_model=ReferenceBookOut)
def rollback_reference(
    book_id: int,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    book = _get_book(book_id, db)
    if not book.is_active:
        raise HTTPException(status_code=422, detail="Справочник не активен")

    book.is_active = False
    book.status = "archived"

    # Re-activate previous consistent version
    prev = (
        db.query(ReferenceBook)
        .filter(
            ReferenceBook.code == book.code,
            ReferenceBook.status == "consistent",
            ReferenceBook.id != book.id,
        )
        .order_by(ReferenceBook.version.desc())
        .first()
    )
    if prev:
        from datetime import datetime
        prev.is_active = True
        prev.activated_at = datetime.utcnow()

    db.add(AuditLog(
        user_id=current_user.id, action="rollback_reference",
        resource_type="reference_book", resource_id=book.id,
    ))
    db.commit()
    db.refresh(book)
    return book


@router.get("/{book_id}/parse")
async def parse_reference(
    book_id: int,
    token: str | None = None,
    db: Session = Depends(get_db),
):
    """SSE stream: parse PDF → reference_rows via Claude vision."""
    from jose import JWTError, jwt as jose_jwt
    from app.config import settings as s
    try:
        payload = jose_jwt.decode(token or "", s.jwt_secret, algorithms=[s.jwt_algorithm])
        user_id = int(payload.get("sub"))
    except JWTError:
        raise HTTPException(status_code=401, detail="Неверный токен")
    user = db.query(User).filter(User.id == user_id).first()
    if not user or user.role != "admin":
        raise HTTPException(status_code=403, detail="Требуются права администратора")

    book = _get_book(book_id, db)
    if not book.pdf_path or not os.path.exists(book.pdf_path):
        raise HTTPException(status_code=422, detail="PDF файл не найден")

    pdf_path = book.pdf_path
    parse_prompt = book.parse_prompt
    book_id_val = book.id
    book_code = book.code

    def _sse(event: str, data: dict) -> str:
        return f"event: {event}\ndata: {json.dumps(data, ensure_ascii=False)}\n\n"

    async def stream() -> AsyncGenerator[str, None]:
        from app.database import SessionLocal
        from app.models import ReferenceRow
        from app.services.reference_parser import parse_reference_pdf

        progress_state = {"page": 0, "total": 0}

        def on_progress(page: int, total: int, msg: str):
            progress_state["page"] = page
            progress_state["total"] = total

        try:
            yield _sse("progress", {"message": "Конвертация PDF в изображения…", "page": 0, "total": 0})
            await asyncio.sleep(0.05)

            loop = asyncio.get_event_loop()
            result = await loop.run_in_executor(
                None,
                lambda: parse_reference_pdf(pdf_path, parse_prompt, on_progress),
            )
            rows, official_name = result

            yield _sse("progress", {"message": f"Сохранение {len(rows)} строк в БД…", "page": progress_state["total"], "total": progress_state["total"]})
            await asyncio.sleep(0.05)

            with SessionLocal() as new_db:
                new_db.query(ReferenceRow).filter(ReferenceRow.book_version_id == book_id_val).delete()
                for r in rows:
                    new_db.add(ReferenceRow(book_version_id=book_id_val, **r))
                book_rec = new_db.query(ReferenceBook).filter(ReferenceBook.id == book_id_val).first()
                if book_rec:
                    book_rec.status = "requires_validation"
                    if official_name and book_rec.official_name == f"СБЦП {book_code}":
                        book_rec.official_name = official_name
                new_db.add(AuditLog(
                    user_id=user_id, action="parse_reference_pdf",
                    resource_type="reference_book", resource_id=book_id_val,
                    details={"rows_parsed": len(rows)},
                ))
                new_db.commit()

            yield _sse("done", {"rows_parsed": len(rows), "book_id": book_id_val})

        except Exception as exc:
            yield _sse("error", {"message": str(exc)})

    return StreamingResponse(stream(), media_type="text/event-stream")


@router.get("/{book_id}/export")
def export_excel(
    book_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    import openpyxl
    book = _get_book(book_id, db)
    rows = db.query(ReferenceRow).filter(ReferenceRow.book_version_id == book_id).order_by(
        ReferenceRow.table_num, ReferenceRow.row_num
    ).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Строки справочника"
    headers = ["table_num", "row_num", "description", "x_min", "x_max", "x_unit", "a", "b", "notes"]
    ws.append(headers)
    for r in rows:
        ws.append([r.table_num, r.row_num, r.description, r.x_min, r.x_max, r.x_unit, r.a, r.b, r.notes])

    # Info sheet
    ws2 = wb.create_sheet("Инфо")
    ws2.append(["Код", book.code])
    ws2.append(["Наименование", book.official_name])
    ws2.append(["Версия", book.version])
    ws2.append(["Строк", len(rows)])
    ws2.append([])
    ws2.append(["Колонки:"])
    ws2.append(["table_num", "Номер таблицы СБЦП (целое)"])
    ws2.append(["row_num", "Номер пункта: п.1, п.2.3 и т.д."])
    ws2.append(["description", "Описание позиции"])
    ws2.append(["x_min", "Минимум X"])
    ws2.append(["x_max", "Максимум X"])
    ws2.append(["x_unit", "Единица X"])
    ws2.append(["a", "Константа формулы a + b*X"])
    ws2.append(["b", "Коэффициент b"])
    ws2.append(["notes", "Примечания"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from urllib.parse import quote
    filename = f"{book.code}_v{book.version}.xlsx"
    encoded = quote(filename, safe="")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


@router.post("/{book_id}/import", response_model=ReferenceBookOut)
def import_excel(
    book_id: int,
    file: UploadFile = File(...),
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    import openpyxl
    book = _get_book(book_id, db)
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=422, detail="Допустимый формат: XLSX")

    content = file.file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    required = {"table_num", "row_num", "description", "a", "b"}
    if not required.issubset(set(headers)):
        raise HTTPException(status_code=422, detail=f"В Excel отсутствуют колонки: {required - set(headers)}")

    col = {h: i for i, h in enumerate(headers)}

    # Delete existing rows for this book version
    db.query(ReferenceRow).filter(ReferenceRow.book_version_id == book_id).delete()

    imported = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        def val(key, default=None):
            idx = col.get(key)
            return row[idx] if idx is not None and idx < len(row) else default

        table_num = val("table_num")
        a_val = val("a")
        b_val = val("b")
        if table_num is None or a_val is None:
            continue

        db.add(ReferenceRow(
            book_version_id=book_id,
            table_num=int(table_num),
            row_num=str(val("row_num", "")) or None,
            description=str(val("description", "")) or None,
            x_min=float(val("x_min")) if val("x_min") is not None else None,
            x_max=float(val("x_max")) if val("x_max") is not None else None,
            x_unit=str(val("x_unit", "")) or None,
            a=float(a_val),
            b=float(b_val) if b_val is not None else None,
            notes=str(val("notes", "")) or None,
        ))
        imported += 1

    book.status = "consistent"
    db.add(AuditLog(
        user_id=current_user.id, action="import_reference_excel",
        resource_type="reference_book", resource_id=book.id,
        details={"rows_imported": imported},
    ))
    db.commit()
    db.refresh(book)
    return book


def _get_book(book_id: int, db: Session) -> ReferenceBook:
    book = db.query(ReferenceBook).filter(ReferenceBook.id == book_id).first()
    if not book:
        raise HTTPException(status_code=404, detail="Справочник не найден")
    return book
